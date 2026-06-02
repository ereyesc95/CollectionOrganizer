"""One-off parser for Collection.xlsx — prints structure and samples."""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

PATH = Path(__file__).parent / "Collection.xlsx"
CATEG_KEYWORDS = (
    "media",
    "autograph",
    "animation",
    "canvas",
    "type",
    "format",
    "edition",
    "signed",
)


def is_green_rgb(rgb) -> bool:
    if rgb is None:
        return False
    if hasattr(rgb, "rgb"):
        rgb = rgb.rgb
    if not isinstance(rgb, str):
        rgb = str(rgb) if rgb else None
    if not rgb or rgb in ("00000000", "FFFFFFFF"):
        return False
    try:
        hex6 = rgb[-6:] if len(rgb) >= 6 else rgb
        r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
        return g > r + 25 and g > b + 15
    except ValueError:
        return False


def main() -> None:
    wb = load_workbook(PATH, data_only=True)
    print("Sheets:", wb.sheetnames)

    for sn in wb.sheetnames:
        ws = wb[sn]
        sep = "=" * 60
        print(f"\n{sep}\nSheet: {sn!r}")
        print(f"max_row={ws.max_row}, max_col={ws.max_column}")

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print("Headers:", headers)
        print(f"Data rows: {max(0, ws.max_row - 1)}")

        for hi, h in enumerate(headers):
            if not h:
                continue
            vals = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, hi + 1).value
                if v is not None and str(v).strip() != "":
                    vals.append(v)
            hl = str(h).lower()
            is_cat = any(k in hl for k in CATEG_KEYWORDS)
            uniq = len({str(v).strip() for v in vals})
            if is_cat or uniq <= 25:
                ctr = Counter(str(v).strip() for v in vals)
                print(f"\n  Column {hi + 1} {h!r}: non_empty={len(vals)}, unique={len(ctr)}")
                for v, c in ctr.most_common(40):
                    print(f"    {c:3d} x {v!r}")
                if len(ctr) > 40:
                    print(f"    ... +{len(ctr) - 40} more")
            else:
                print(
                    f"\n  Column {hi + 1} {h!r}: non_empty={len(vals)}, unique={uniq} (text)"
                )
                samples = list(dict.fromkeys(str(v).strip() for v in vals))[:8]
                for s in samples:
                    tail = "..." if len(s) > 100 else ""
                    print(f"    sample: {s[:100]!r}{tail}")

        album_col = next(
            (i for i, h in enumerate(headers) if h and "album" in str(h).lower()),
            None,
        )
        if album_col is not None:
            albums = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, album_col + 1).value
                if v:
                    albums.append(str(v).strip())
            print(f"\n  ALBUM column: {len(albums)} entries")
            simple, edition_comma, other = [], [], []
            for a in albums:
                parts = a.split(",", 1)
                if len(parts) < 2:
                    other.append(a)
                    continue
                after_first = parts[1]
                dot_pos = after_first.find(".")
                tail = after_first[dot_pos + 1 :] if dot_pos >= 0 else ""
                if re.search(r"\d{4}\.", after_first) and re.search(r",\s*\d{4}", tail):
                    edition_comma.append(a)
                elif re.search(r"\d{4}\.", after_first):
                    simple.append(a)
                else:
                    other.append(a)
            print(
                f"  Pattern counts: simple={len(simple)}, "
                f"edition_comma={len(edition_comma)}, other={len(other)}"
            )
            print("  Edition-style examples (up to 12):")
            for e in edition_comma[:12]:
                print(f"    {e}")
            print("  Other/nonstandard (all):")
            for e in other:
                print(f"    {e}")

        green_cols: dict[str, list[int]] = {}
        for c in range(1, ws.max_column + 1):
            h = headers[c - 1] if c <= len(headers) else ""
            green_rows: list[int] = []
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, c)
                fill = cell.fill
                if fill and fill.fill_type == "solid":
                    fg = fill.fgColor
                    rgb = getattr(fg, "rgb", None) if fg else None
                    if is_green_rgb(rgb) and len(green_rows) < 5:
                        green_rows.append(r)
            if green_rows:
                green_cols[str(h)] = green_rows
        if green_cols:
            print("\n  Columns with green fill (sample rows):", green_cols)
        else:
            print("\n  No green fill detected")

    wb.close()


if __name__ == "__main__":
    main()
