from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.crud import create_record
from app.models import Record
from app.parse_album import parse_album, split_tags
from app.schemas import ImportResult, RecordCreate


def import_from_xlsx(db: Session, path: Path, *, replace: bool = False) -> ImportResult:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers):
            if h == name.upper():
                return i + 1
        return None

    album_c = col("ALBUM")
    anim_c = col("ANIMATION")
    canvas_c = col("CANVAS")
    media_c = col("MEDIA TYPE")
    auto_c = col("AUTOGRAPHS")
    pending_c = col("PENDING")

    if not album_c:
        return ImportResult(imported=0, updated=0, skipped=0, errors=["ALBUM column not found"])

    if replace:
        db.query(Record).delete()
        db.commit()

    imported = updated = skipped = 0
    errors: list[str] = []

    for row in range(2, ws.max_row + 1):
        album = ws.cell(row, album_c).value
        if not album or not str(album).strip():
            skipped += 1
            continue

        cover_key = str(album).strip()
        parsed = parse_album(cover_key)

        def cell(c: int | None) -> str | None:
            if not c:
                return None
            v = ws.cell(row, c).value
            return str(v).strip() if v is not None and str(v).strip() else None

        media_tags = split_tags(cell(media_c))

        animation_tags = split_tags(cell(anim_c))
        canvas_tags = split_tags(cell(canvas_c))
        autograph_tags = split_tags(cell(auto_c))
        pending = cell(pending_c)

        existing = db.query(Record).filter(Record.cover_key == cover_key).first()
        if existing:
            existing.artist = parsed.artist
            existing.record_year = parsed.record_year
            existing.title = parsed.title
            existing.edition_year = parsed.edition_year
            existing.edition_title = parsed.edition_title
            existing.pending = pending
            from app.crud import _set_tags
            from app.models import (
                RecordAnimationTag,
                RecordAutographTag,
                RecordCanvasTag,
                RecordMediaTag,
            )

            _set_tags(db, existing, RecordMediaTag, "media_tags", media_tags)
            _set_tags(db, existing, RecordAnimationTag, "animation_tags", animation_tags)
            _set_tags(db, existing, RecordCanvasTag, "canvas_tags", canvas_tags)
            _set_tags(db, existing, RecordAutographTag, "autograph_tags", autograph_tags)
            db.commit()
            updated += 1
            continue

        try:
            create_record(
                db,
                RecordCreate(
                    cover_key=cover_key,
                    artist=parsed.artist,
                    record_year=parsed.record_year,
                    title=parsed.title,
                    edition_year=parsed.edition_year,
                    edition_title=parsed.edition_title,
                    pending=pending,
                    media_tags=media_tags,
                    animation_tags=animation_tags,
                    canvas_tags=canvas_tags,
                    autograph_tags=autograph_tags,
                ),
            )
            imported += 1
        except Exception as e:
            errors.append(f"Row {row}: {e}")
            skipped += 1

    wb.close()
    return ImportResult(imported=imported, updated=updated, skipped=skipped, errors=errors[:20])
